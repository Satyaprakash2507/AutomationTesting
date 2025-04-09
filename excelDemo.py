import openpyxl

book = openpyxl.load_workbook('C:\\Users\\yaduv\\Documents\\excelDemo.xlsx')

sheet = book.active  # Get the active sheet

Dict = {}  # Initialize an empty dictionary

cell = sheet.cell(row=1, column=2)  # Get the cell at row 1, column 2
print(cell.value)

sheet.cell(row=2, column=2).value = 'Hello World'  # Set the value of the cell at row 2, column 2

print(sheet.cell(row=2, column=2).value)  # Print the updated value of the cell

print(sheet.max_row)  # Print the maximum number of rows in the sheet
print(sheet.max_column)  # Print the maximum number of columns in the sheet

#print all the value of column1

#for i in range(1, sheet.max_row + 1):
 #   print(sheet.cell(row=i, column=1).value)  

 #print all the values of row and column

#for i in range(1, sheet.max_row + 1):
#    for j in range(1, sheet.max_column + 1):
#        print(sheet.cell(row=i, column=j).value)  


#print the vale of testcase2 only
for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == 'testcase2':
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row=i, column=j).value)  


#adding the value in the dictionary

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=1).value == 'testcase2':
        for j in range(2, sheet.max_column + 1):
           Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value #Dict["lastname"] = "yadav"
print(Dict)  