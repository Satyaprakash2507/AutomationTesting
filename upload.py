import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


def update_excel_data(file_path, searchTerm,colName,new_value):
    book = openpyxl.load_workbook(file_path)
    sheet = book.active  # Get the active sheet
    
    Dict = {}  # Initialize an empty dictionary
# Get the column index of the 'price' column
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            Dict["col"] = i
# Get the row index of the 'Apple' row

    for j in range(1, sheet.max_row + 1):
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column= j).value == searchTerm:
                Dict["row"] = i

#edit the excel with updated values.
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = 1000  
    book.save(file_path)  # Save the workbook with the updated value










file_path = "C:\\Users\\yaduv\\Downloads\\download.xlsx"
fruit_name = "Apple"
new_value = 1000

driver = webdriver.Chrome()
driver.implicitly_wait(5)  


driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

driver.find_element(By.ID, "downloadButton").click()  

#calling the function to update the excel file with new values.
update_excel_data(file_path,fruit_name,"price",new_value)


#upload the file
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 10)
toast_locator = (By.CSS_SELECTOR,".Toastify__toast-body div:nth-child(2)")
wait.until(EC.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)


priceColumn = driver.find_element(By.XPATH,"//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text
print(actual_price)


assert actual_price == str(new_value)